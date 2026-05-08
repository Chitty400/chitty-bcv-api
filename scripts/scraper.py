"""
BCV Scraper - Obtiene tasa de cambio USD e IPC del Banco Central de Venezuela
Diseñado para ejecutarse en GitHub Actions y publicar JSON estático en GitHub Pages
"""

import json
import os
import re
import sys
import urllib.request
import urllib.error
from datetime import datetime, timezone
from pathlib import Path

# ── Dependencias opcionales (se instalan en el workflow) ──────────────────────
try:
    from bs4 import BeautifulSoup
    import openpyxl
    from openpyxl import load_workbook
    import io
except ImportError:
    print("ERROR: Instala dependencias: pip install beautifulsoup4 openpyxl requests lxml")
    sys.exit(1)

try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False

# ── Constantes ────────────────────────────────────────────────────────────────
BCV_HOME_URL     = "https://www.bcv.org.ve/"
BCV_IPC_XLS_URL  = (
    "https://www.bcv.org.ve/sites/default/files/precios_consumidor/"
    "4_5_7_indice_y_variaciones_mensuales_serie_desde_dic_2007_1.xls"
)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; BCV-Static-API/1.0; "
        "+https://github.com/TU_USUARIO/bcv-api)"
    ),
    "Accept-Language": "es-VE,es;q=0.9,en;q=0.8",
}

OUTPUT_DIR   = Path(__file__).parent.parent / "docs"
LATEST_FILE  = OUTPUT_DIR / "latest.json"
HISTORY_FILE = OUTPUT_DIR / "history.json"


# ── Helpers HTTP ──────────────────────────────────────────────────────────────

def http_get_text(url: str) -> str:
    """GET con requests o urllib como fallback."""
    if REQUESTS_AVAILABLE:
        r = requests.get(url, headers=HEADERS, timeout=30, verify=False)
        r.raise_for_status()
        return r.text
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read().decode("utf-8", errors="replace")


def http_get_bytes(url: str) -> bytes:
    """GET binario (para el XLS)."""
    if REQUESTS_AVAILABLE:
        r = requests.get(url, headers=HEADERS, timeout=60, verify=False)
        r.raise_for_status()
        return r.content
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


# ── Scraper de Tasa de Cambio ────────────────────────────────────────────────

def scrape_tasa_bcv() -> float:
    """
    Extrae la tasa USD/VES de la página principal del BCV.
    El BCV publica el valor dentro de un <div id="dolar"> o similar.
    Estrategia: buscar el valor numérico con formato 'X.XXXXX' cerca de 'USD' o 'dólar'.
    """
    html = http_get_text(BCV_HOME_URL)
    soup = BeautifulSoup(html, "lxml")

    # Estrategia 1: div con id o clase relacionada al dólar
    for selector in ["#dolar strong", ".dolar strong", "#dolar", ".tipo-cambio-dolar"]:
        tag = soup.select_one(selector)
        if tag:
            text = tag.get_text(strip=True).replace(",", ".")
            m = re.search(r"\d{2,6}[.,]\d{2,8}", text)
            if m:
                return float(m.group().replace(",", "."))

    # Estrategia 2: buscar el patrón numérico grande más cercano a "USD" en toda la página
    # El BCV muestra algo como "569,42730764" cerca del label dólar
    text_full = soup.get_text(" ")
    # Buscar secuencias tipo 569,42730764 o 100.26000000
    candidates = re.findall(r"\b\d{2,6}[,\.]\d{5,10}\b", text_full)
    if candidates:
        # El primero suele ser el USD en la página principal
        return float(candidates[0].replace(",", "."))

    raise ValueError("No se pudo extraer la tasa de cambio del BCV")


# ── Scraper de IPC ────────────────────────────────────────────────────────────

def _parse_xls_bytes(content: bytes) -> dict:
    """
    Parsea el XLS del BCV con openpyxl.
    Columnas esperadas: Fecha | Índice | Variación mensual % | Variación anual %
    El archivo tiene encabezados en las primeras filas; los datos empiezan después.
    Retorna el registro más reciente.
    """
    # openpyxl no soporta .xls (formato antiguo) — usamos xlrd si está disponible
    try:
        import xlrd
        book = xlrd.open_workbook(file_contents=content)
        sheet = book.sheet_by_index(0)

        # Buscar la última fila con datos (de abajo hacia arriba)
        last_row = None
        for i in range(sheet.nrows - 1, -1, -1):
            row = sheet.row_values(i)
            # Fila válida: primera celda es texto de fecha o año, tercera/cuarta son números
            if row[0] and row[2] and isinstance(row[2], (int, float)) and float(row[2]) != 0:
                last_row = row
                break

        if last_row is None:
            raise ValueError("No se encontraron filas de datos en el XLS")

        fecha_raw = str(last_row[0]).strip()
        var_mensual = float(last_row[2])
        var_anual   = float(last_row[3]) if len(last_row) > 3 and last_row[3] else None

        # Normalizar fecha → "YYYY-MM"
        fecha = _normalize_fecha(fecha_raw)
        return {
            "fecha": fecha,
            "variacion_mensual": round(var_mensual, 2),
            "variacion_anual": round(var_anual, 2) if var_anual else None,
        }

    except ImportError:
        pass

    # Fallback: intentar con openpyxl (puede fallar con .xls antiguo)
    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        last_row = None
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None and row[2] is not None:
                try:
                    float(row[2])
                    last_row = row
                except (TypeError, ValueError):
                    pass

        if last_row is None:
            raise ValueError("No se encontraron datos en el XLS (openpyxl)")

        fecha   = _normalize_fecha(str(last_row[0]))
        var_m   = float(last_row[2])
        var_a   = float(last_row[3]) if len(last_row) > 3 and last_row[3] is not None else None

        return {
            "fecha": fecha,
            "variacion_mensual": round(var_m, 2),
            "variacion_anual": round(var_a, 2) if var_a else None,
        }
    except Exception as e:
        raise ValueError(f"No se pudo parsear el XLS: {e}")


def _normalize_fecha(raw: str) -> str:
    """
    Intenta parsear distintos formatos de fecha que usa el BCV
    y devuelve 'YYYY-MM'.
    Ejemplos de entrada: 'ene-26', 'Enero 2026', '2026-01', '01/2026'
    """
    raw = raw.strip()

    # Ya está en formato YYYY-MM
    if re.match(r"^\d{4}-\d{2}$", raw):
        return raw

    # Formato dd/mm/aaaa o mm/aaaa
    m = re.match(r"^(\d{1,2})/(\d{4})$", raw)
    if m:
        return f"{m.group(2)}-{int(m.group(1)):02d}"

    # Nombre de mes en español abreviado o completo + año de 2 o 4 dígitos
    MESES = {
        "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
        "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
        "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5,
        "junio": 6, "julio": 7, "agosto": 8, "septiembre": 9,
        "octubre": 10, "noviembre": 11, "diciembre": 12,
    }
    m = re.match(r"^([a-záéíóú]+)[.\-\s]+(\d{2,4})$", raw.lower())
    if m:
        mes_str, anio = m.group(1), m.group(2)
        mes_num = MESES.get(mes_str[:3])
        if mes_num:
            anio_full = int(anio) + 2000 if len(anio) == 2 else int(anio)
            return f"{anio_full}-{mes_num:02d}"

    # Año + mes abreviado: "2026 ene"
    m = re.match(r"^(\d{4})\s+([a-z]+)$", raw.lower())
    if m:
        mes_num = MESES.get(m.group(2)[:3])
        if mes_num:
            return f"{m.group(1)}-{mes_num:02d}"

    # Último recurso: devolver tal cual con advertencia
    print(f"  ADVERTENCIA: No se pudo normalizar la fecha '{raw}', se usa literal")
    return raw


def scrape_ipc_bcv() -> dict:
    """Descarga el XLS del IPC del BCV y devuelve el dato más reciente."""
    print(f"  Descargando XLS IPC desde: {BCV_IPC_XLS_URL}")
    content = http_get_bytes(BCV_IPC_XLS_URL)
    print(f"  XLS descargado: {len(content):,} bytes")
    return _parse_xls_bytes(content)


# ── Gestión de archivos JSON ──────────────────────────────────────────────────

def load_json(path: Path) -> dict | list | None:
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def save_json(path: Path, data, indent: int = 2):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=indent)
    print(f"  Guardado: {path}")


# ── Lógica principal ──────────────────────────────────────────────────────────

def build_latest(tasa: float, ipc: dict) -> dict:
    return {
        "updated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "tasa_bcv": tasa,
        "ipc": ipc,
    }


def update_history(history_data: list | None, latest: dict) -> list:
    """Añade el registro actual al historial si es nuevo."""
    if history_data is None:
        history_data = []

    # Clave única: fecha del IPC + fecha del día para la tasa
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    ipc_fecha = latest["ipc"]["fecha"]

    # Evitar duplicados: si ya existe entrada del mismo día, actualizar
    for i, entry in enumerate(history_data):
        if entry.get("date") == today:
            history_data[i] = {"date": today, **latest}
            print(f"  Historial: actualizada entrada del {today}")
            return history_data

    history_data.append({"date": today, **latest})
    print(f"  Historial: añadida nueva entrada para {today} (IPC: {ipc_fecha})")
    return history_data


def main():
    print("=" * 60)
    print("BCV Static API - Scraper")
    print("=" * 60)

    # Determinar qué actualizar según variable de entorno o argumento
    mode = os.environ.get("SCRAPER_MODE", "tasa").lower()
    if len(sys.argv) > 1:
        mode = sys.argv[1].lower()

    print(f"Modo: {mode}")

    # Cargar datos previos
    latest_prev  = load_json(LATEST_FILE) or {}
    history_data = load_json(HISTORY_FILE)

    tasa = latest_prev.get("tasa_bcv")
    ipc  = latest_prev.get("ipc", {})

    # ── Obtener Tasa ──────────────────────────────────────────────────────
    if mode in ("tasa", "all", "both"):
        print("\n[1/2] Scrapeando tasa de cambio BCV...")
        try:
            tasa = scrape_tasa_bcv()
            print(f"  ✓ Tasa USD/VES: {tasa}")
        except Exception as e:
            print(f"  ✗ ERROR obteniendo tasa: {e}")
            if tasa is None:
                print("  No hay tasa previa disponible, abortando.")
                sys.exit(1)
            print(f"  Usando tasa previa: {tasa}")

    # ── Obtener IPC ───────────────────────────────────────────────────────
    if mode in ("ipc", "all", "both"):
        print("\n[2/2] Descargando IPC del BCV...")
        try:
            ipc = scrape_ipc_bcv()
            print(f"  ✓ IPC: {ipc}")
        except Exception as e:
            print(f"  ✗ ERROR obteniendo IPC: {e}")
            if not ipc:
                print("  No hay IPC previo disponible, abortando.")
                sys.exit(1)
            print(f"  Usando IPC previo: {ipc}")

    # ── Construir y guardar ───────────────────────────────────────────────
    if tasa is None or not ipc:
        print("\nERROR: Datos incompletos, no se puede generar JSON.")
        sys.exit(1)

    latest = build_latest(tasa, ipc)
    history_data = update_history(history_data, latest)

    print("\nGuardando archivos...")
    save_json(LATEST_FILE, latest)
    save_json(HISTORY_FILE, history_data)

    print("\n✓ Completado exitosamente")
    print(f"  Tasa: {latest['tasa_bcv']}")
    print(f"  IPC:  {latest['ipc']}")
    print(f"  URL latest: https://TU_USUARIO.github.io/bcv-api/latest.json")


if __name__ == "__main__":
    main()
